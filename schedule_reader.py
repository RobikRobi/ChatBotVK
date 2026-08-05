"""
Модуль для загрузки расписания из Google Таблиц или xlsx файла.

Для использования Google Таблиц необходимо:
1. Создать сервисный аккаунт в Google Cloud Console
2. Скачать JSON-файл с ключами (credentials.json)
3. Предоставить доступ к Google Таблице email сервисного аккаунта
4. Указать путь к credentials.json и ID таблицы в .env файле
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional


ROOM_RE = re.compile(r"\s*\d+(?:[\\/]\d+)*$")
DAY_NAMES = ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница")


@dataclass(frozen=True)
class Lesson:
    day: str
    time: str
    class_name: str
    subject: str


@dataclass(frozen=True)
class ScheduleData:
    classes: list[str]
    grades: dict[str, list[str]]
    subjects: list[str]
    lessons: list[Lesson]

    def lessons_for_class(self, class_name: str) -> list[Lesson]:
        return [lesson for lesson in self.lessons if lesson.class_name == class_name and lesson.subject]

    def lessons_for_teacher(self, subjects: list[str], classes: list[str]) -> list[Lesson]:
        subject_set = set(subjects)
        class_set = set(classes)
        result = []
        for lesson in self.lessons:
            if lesson.class_name not in class_set or not lesson.subject:
                continue
            lesson_subjects = subjects_from_cell(lesson.subject)
            if subject_set.intersection(lesson_subjects):
                result.append(lesson)
        return result


def _value(cell):
    return "" if cell is None else str(cell).strip()


def normalize_subject(text: str) -> str:
    text = " ".join(text.replace("\\\n", " ").replace("\n", " ").split())
    replacements = {
        "Английский яз.": "Английский язык",
        "Индивид.проект": "Индивидуальный проект",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return ROOM_RE.sub("", text).strip()


def subjects_from_cell(text: str) -> list[str]:
    subjects = []
    for part in re.split(r"\s*/\s*|\s*\\\s*", text):
        subject = normalize_subject(part)
        if subject and not subject.isdigit():
            subjects.append(subject)
    return subjects


def _build_grades(classes: list[str]) -> dict[str, list[str]]:
    grades: dict[str, list[str]] = {}
    for class_name in classes:
        grade = re.match(r"\d+", class_name)
        if grade:
            grades.setdefault(grade.group(0), []).append(class_name)
    return {
        grade: sorted(names)
        for grade, names in sorted(grades.items(), key=lambda item: int(item[0]))
    }


def parse_google_sheet(worksheet) -> ScheduleData:
    """
    Парсит данные из листа Google Таблицы.
    
    Ожидается структура:
    - Строка 1: заголовки с классами (начиная с колонки C)
    - Колонка A: время уроков
    - Колонка B: пустая или разделитель дней
    - Колонки C+: предметы по классам
    
    Для разделения по дням недели можно использовать:
    - Пустые строки между днями
    - Или отдельный лист на каждый день
    """
    all_values = worksheet.get_all_values()
    
    if not all_values:
        raise RuntimeError("Лист таблицы пустой.")
    
    # Находим заголовок с классами (первая строка, где во второй колонке "Звонки" или есть время)
    header_index = 0
    for idx, row in enumerate(all_values):
        if len(row) > 1 and (row[1] == "Звонки" or (len(row[1]) >= 3 and ":" in row[1])):
            header_index = idx
            break
    else:
        # Если не нашли явный маркер, считаем первую строку заголовком
        header_index = 0
    
    classes = [name for name in all_values[header_index][2:] if name]
    class_set = set(classes)
    subjects = set()
    lessons: list[Lesson] = []
    
    # Определяем блоки дней (разделены пустыми строками или строками с названием дня)
    current_day = "Понедельник"
    day_counter = 0
    
    for row_idx in range(header_index + 1, len(all_values)):
        row = all_values[row_idx]
        
        # Пропускаем пустые строки
        if not row or all(not cell for cell in row):
            # Пустая строка может означать переход к следующему дню
            day_counter += 1
            if day_counter < len(DAY_NAMES):
                current_day = DAY_NAMES[day_counter]
            continue
        
        # Проверяем, не является ли строка заголовком нового дня
        if len(row) > 0 and row[0] in DAY_NAMES:
            current_day = row[0]
            continue
        
        # Пропускаем строки без времени
        if len(row) < 2 or not row[1]:
            continue
        
        lesson_time = row[1]
        
        # Обрабатываем ячейки с предметами
        for col_index, class_name in enumerate(all_values[header_index][2:], start=2):
            if not class_name or class_name not in class_set:
                continue
            
            cell_text = row[col_index] if col_index < len(row) else ""
            
            if cell_text:
                subjects.update(subjects_from_cell(cell_text))
            
            lessons.append(
                Lesson(
                    day=current_day,
                    time=lesson_time,
                    class_name=class_name,
                    subject=cell_text,
                )
            )
    
    return ScheduleData(
        classes=classes,
        grades=_build_grades(classes),
        subjects=sorted(subjects),
        lessons=lessons,
    )


def load_schedule_from_google(
    credentials_path: Path,
    spreadsheet_id: str,
    worksheet_name: Optional[str] = None
) -> ScheduleData:
    """
    Загружает расписание из Google Таблицы.
    
    Args:
        credentials_path: Путь к JSON-файлу с ключами сервисного аккаунта
        spreadsheet_id: ID Google Таблицы (из URL: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit)
        worksheet_name: Имя листа (по умолчанию первый лист)
    
    Returns:
        ScheduleData: Объект с данными расписания
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as e:
        raise ImportError(
            "Не установлены библиотеки для работы с Google Sheets. "
            "Выполните: pip install gspread google-auth google-auth-oauthlib google-auth-httplib2"
        ) from e
    
    if not credentials_path.exists():
        raise FileNotFoundError(f"Файл с ключами не найден: {credentials_path}")
    
    # Создаем credentials из JSON-файла
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly"
    ]
    creds = Credentials.from_service_account_file(
        str(credentials_path),
        scopes=scopes
    )
    
    # Подключаемся к Google Sheets API
    client = gspread.authorize(creds)
    
    # Открываем таблицу по ID
    spreadsheet = client.open_by_key(spreadsheet_id)
    
    # Выбираем лист
    if worksheet_name:
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            raise ValueError(f"Лист '{worksheet_name}' не найден в таблице")
    else:
        worksheet = spreadsheet.sheet1
    
    return parse_google_sheet(worksheet)


# Для совместимости со старым кодом
def load_schedule(path: Path) -> ScheduleData:
    """
    Загружает расписание из файла (xlsx) или Google Таблицы.
    
    Если путь выглядит как ID Google Таблицы (содержит 'docs.google.com' или длинную строку),
    пытается загрузить из Google Sheets. Иначе загружает из xlsx файла.
    """
    path_str = str(path)
    
    # Проверяем, не является ли это ссылкой на Google Таблицу или ID
    if "docs.google.com" in path_str or (len(path_str) > 20 and not path_str.endswith(('.xlsx', '.xls'))):
        # Это ID таблицы или ссылка
        credentials_path = Path(__file__).parent / "credentials.json"
        
        # Извлекаем ID из ссылки если нужно
        if "docs.google.com" in path_str:
            match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', path_str)
            if match:
                spreadsheet_id = match.group(1)
            else:
                raise ValueError("Не удалось извлечь ID таблицы из ссылки")
        else:
            spreadsheet_id = path_str
        
        return load_schedule_from_google(credentials_path, spreadsheet_id)
    
    # Загружаем из xlsx файла (старый метод)
    from openpyxl import load_workbook
    
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = [[_value(cell.value) for cell in row] for row in sheet.iter_rows()]

    header_indexes = [
        index for index, row in enumerate(rows)
        if len(row) > 2 and row[1] == "Звонки"
    ]
    if not header_indexes:
        raise RuntimeError("Не удалось найти строку с классами в расписании.")

    classes = [name for name in rows[header_indexes[0]][2:] if name]
    class_set = set(classes)
    subjects = set()
    lessons: list[Lesson] = []

    for block_number, header_index in enumerate(header_indexes):
        day = DAY_NAMES[block_number] if block_number < len(DAY_NAMES) else f"День {block_number + 1}"
        header = rows[header_index]
        next_header = header_indexes[block_number + 1] if block_number + 1 < len(header_indexes) else len(rows)

        for row in rows[header_index + 1:next_header]:
            if len(row) < 3 or not row[1]:
                continue
            lesson_time = row[1]
            for col_index, class_name in enumerate(header[2:], start=2):
                if not class_name or class_name not in class_set:
                    continue
                cell_text = row[col_index] if col_index < len(row) else ""
                if cell_text:
                    subjects.update(subjects_from_cell(cell_text))
                lessons.append(
                    Lesson(
                        day=day,
                        time=lesson_time,
                        class_name=class_name,
                        subject=cell_text,
                    )
                )

    return ScheduleData(
        classes=classes,
        grades=_build_grades(classes),
        subjects=sorted(subjects),
        lessons=lessons,
    )
